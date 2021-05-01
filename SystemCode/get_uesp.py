import os
import rpa as r
import pandas as pd
from pandas import DataFrame
from datetime import datetime
from openpyxl import load_workbook
from sklearn import linear_model
import statsmodels.api as sm
from sklearn.preprocessing import StandardScaler  
from sklearn.preprocessing import PolynomialFeatures
from sklearn.linear_model import LinearRegression
from sklearn.pipeline import Pipeline
import matplotlib.pyplot as plt
from sklearn.metrics import mean_squared_error, r2_score
import numpy as np

def get_period_nbr():
    now = datetime.now()
    if (now.minute>30):
        period_nbr = now.hour*2+2
    else:
        period_nbr = now.hour*2+1
    return period_nbr

def append_df_to_excel(filename, df, sheet_name='DATA', startrow=None,truncate_sheet=False,**to_excel_kwargs):
    
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    writer.book = load_workbook(filename)
    
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

def predict_usep(df,period):
	#df = pd.read_excel(filename,'Sheet1')
	X = df[['PERIOD']]
	# X = df[['DEMAND']]
	y = df[['DEMAND']]
	z = df[['USEP']]

	model_demand = Pipeline([('poly', PolynomialFeatures(degree=6)),
	              ('linear', LinearRegression(fit_intercept=False))])

	demand_lr=model_demand.fit(X,y) 

	p=period

	demand_predicted_arr = demand_lr.predict(np.array([[p]]))

	demand_predicted = demand_predicted_arr[0][0]

	model_usep = Pipeline([('poly', PolynomialFeatures(degree=5)),
              ('linear', LinearRegression(fit_intercept=False))])

	usep_lr=model_usep.fit(y,z) 

	usep_pred_arr = usep_lr.predict(np.array([[demand_predicted]]))

	usep_pred =usep_pred_arr[0][0]

	return demand_predicted,usep_pred

print(get_period_nbr())
period_nbr=get_period_nbr()
today = datetime.today()
#os.chdir("D:")
r.init()
r.url('https://www.emcsg.com/marketdata/priceinformation')
r.wait(2)
print(r.read('//table[@class="ptable realtimePriceTable"]//tr[@class="current"]'))
data_list=r.read('//table[@class="ptable realtimePriceTable"]//tr[@class="current"]').splitlines()
data_list[0]=today.strftime('%d/%m/%Y')
data_list[1]=get_period_nbr()
data_list[2]=float(data_list[2])
data_list[3]=float(data_list[3])
data_list[4]=float(data_list[4])
data_list[6]=float(data_list[6])
if data_list[6]==0:
    data_list.append(0)
else:
    data_list.append(1)
data_list=[data_list]
df_current = DataFrame (data_list,columns=['Date','Period','Demand','TCL','USEP','EHEUR','LCP','Regulation','Primary','Secondary','Contingency','DR'])
df_current = df_current[['Date','Period','USEP','LCP','Demand','TCL','DR']]
filename = '/Users/chengwen/Desktop/ISA/Project/DemandForecastData.xlsx'
df = pd.read_excel(filename,'Sheet1')
if df.iloc[-1]['PERIOD']!=period_nbr:
    print ("append")
    append_df_to_excel(filename, df_current, sheet_name='Sheet1', header=None,index=False)
r.close()
demand_predicted1,usep_pred1=predict_usep(df,period_nbr+3)
demand_predicted2,usep_pred2=predict_usep(df,period_nbr+4)
demand_predicted3,usep_pred3=predict_usep(df,period_nbr+5)
demand_predicted4,usep_pred4=predict_usep(df,period_nbr+6)
demand_predicted5,usep_pred5=predict_usep(df,period_nbr+7)
demand_usep = [[period_nbr+3,demand_predicted1,usep_pred1],[period_nbr+4,demand_predicted2,usep_pred2],[period_nbr+5,demand_predicted3,usep_pred3],[period_nbr+6,demand_predicted4,usep_pred4],[period_nbr+7,demand_predicted5,usep_pred5]]
df = DataFrame (demand_usep,columns=['Period','PredictedDemand','PredictedUSEP'])
print (df)
